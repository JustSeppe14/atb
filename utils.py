import pandas as pd
import os
from openpyxl import load_workbook
import shutil
from datetime import datetime


DEELNEMERS_FILE = "Deelnemers/deelnemerslijst 2025.xlsx"
RESULT_FILE = "Result/finish.xlsx"
TEMPLATE_FILE = "Template/klassement.xlsx"
MAX_POINTS = 80
_CURRENT_BACKUP_DIR = None

def load_deelnemers():
    df = pd.read_excel(DEELNEMERS_FILE, header=4)
    df.columns = df.columns.str.strip().str.lower()
    df = df.rename(columns={
        'number': 'bib',
        'name': 'naam',
        'klasse': 'klasse',
        'cat': 'categorie',
        'team': 'team'
    })
    df = df.dropna(subset=['bib', 'naam', 'klasse'])
    df['naam'] = df['naam'].str.strip().str.lower()
    df['bib'] = df['bib'].astype(int)
    return df

def load_result():
    df = pd.read_excel(RESULT_FILE)
    df.columns = df.columns.str.strip().str.lower()
    df = df.rename(columns={'pl': 'plaats', 'bib': 'bib', 'naam': 'naam'})
    df['bib'] = pd.to_numeric(df['bib'], errors='coerce').astype('Int64')
    df = df.dropna(subset=['bib', 'plaats'])
    return df

def get_current_week(overall_path, sheet_name):
    # If file doesn't exist, don't create anything — just start at week 1
    if not os.path.isfile(overall_path):
        return 1

    wb = load_workbook(overall_path)

    # Create the sheet if it's missing
    if sheet_name not in wb.sheetnames:
        wb.create_sheet(title=sheet_name)
        wb.save(overall_path)
        return 1

    # Sheet exists — check existing week columns
    df = pd.read_excel(overall_path, sheet_name=sheet_name)
    week_cols = [col for col in df.columns if str(col).isdigit()]
    return len(week_cols) + 1

def load_template_column_order():
    template_df = pd.read_excel(TEMPLATE_FILE, sheet_name=0, nrows=0)
    return [col for col in template_df.columns if not str(col).startswith("Unnamed")]

def backup_deelnemers_file():
    """
    Backup the deelnemers file to a new file with a timestamp in the filename.
    """
    backup_dir = "Deelnemers/backups"
    os.makedirs(backup_dir, exist_ok=True)
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = os.path.join(backup_dir, f"deelnemerslijst_2025_{timestamp}.xlsx")
    shutil.copy(DEELNEMERS_FILE, backup_path)
    
def detect_klasse_wissels_met_backup():
    backup_dir = "Deelnemers/backups"
    if not os.path.exists(backup_dir):
        return {}
    
    # Zoek de meest recente backup
    backups = [f for f in os.listdir(backup_dir) if f.endswith('.xlsx')]
    if not backups:
        return {}
    
    latest_backup = max(backups, key=lambda f: os.path.getmtime(os.path.join(backup_dir, f)))
    backup_path = os.path.join(backup_dir, latest_backup)
    
     # Laad beide lijsten
    df_now = pd.read_excel(DEELNEMERS_FILE, header=4)
    df_now.columns = df_now.columns.str.strip().str.lower()
    df_now = df_now.rename(columns={'number': 'bib', 'klasse': 'klasse'})
    df_now = df_now.dropna(subset=['bib', 'klasse'])
    df_now['bib'] = df_now['bib'].astype(int)

    df_old = pd.read_excel(backup_path, header=4)
    df_old.columns = df_old.columns.str.strip().str.lower()
    df_old = df_old.rename(columns={'number': 'bib', 'klasse': 'klasse'})
    df_old = df_old.dropna(subset=['bib', 'klasse'])
    df_old['bib'] = df_old['bib'].astype(int)

    # Vergelijk klasse per bib
    merged = pd.merge(df_old[['bib', 'klasse']], df_now[['bib', 'klasse']], on='bib', suffixes=('_oud', '_nieuw'))
    wissels = merged[merged['klasse_oud'] != merged['klasse_nieuw']]
    return {row['bib']: (row['klasse_oud'], row['klasse_nieuw']) for _, row in wissels.iterrows()}

def get_current_backup_dir():
    """Get or create the backup directory for the current run."""
    global _CURRENT_BACKUP_DIR
    if _CURRENT_BACKUP_DIR is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        _CURRENT_BACKUP_DIR = os.path.join("output_backups", timestamp)
        os.makedirs(_CURRENT_BACKUP_DIR, exist_ok=True)
    return _CURRENT_BACKUP_DIR

def backup_file(source_file, backup_name=None):
    """Backup a file to the current run's backup directory."""
    backup_dir = get_current_backup_dir()
    if backup_name is None:
        backup_name = os.path.basename(source_file)
    backup_file_path = os.path.join(backup_dir, backup_name)
    shutil.copy2(source_file, backup_file_path)
    return backup_file_path

    