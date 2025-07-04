import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import logging
import shutil
from datetime import datetime

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)
from utils import (
    load_deelnemers,
    load_result,
    get_current_week,
    load_template_column_order,
    detect_klasse_wissels_met_backup,
    backup_file,
    MAX_POINTS,
    DEELNEMERS_FILE,
    RESULT_FILE,
    TEMPLATE_FILE
)
OUTPUT_DIR = "output"
os.makedirs(OUTPUT_DIR, exist_ok=True)
KLASSEMENT_FILE = os.path.join(OUTPUT_DIR, "klassement_2025.xlsx")
IS_SECOND_PERIOD_STARTED = os.environ.get('IS_SECOND_PERIOD_STARTED', 'False').lower() == 'true'

def generate_regelmatigheidscriterium():
    try:
        deelnemers = load_deelnemers()
        uitslag = load_result()
        week_num = get_current_week(KLASSEMENT_FILE, sheet_name="REGELMATIGHEIDSCRITERIUM")
        week_col = str(week_num)
        
        logger.info(f"Generating regelmatigheidscriterium for week {week_col}")

        punten_per_rijder = []
        klasse_per_rijder = []
        categorie_per_rijder = []

        for klasse in deelnemers['klasse'].unique():
            klasse_deelnemers = deelnemers[deelnemers['klasse'] == klasse]
            klasse_bibs = set(klasse_deelnemers['bib'])

            klasse_result = uitslag[uitslag['bib'].isin(klasse_bibs)].copy()
            klasse_result['rank_in_class'] = klasse_result['plaats'].rank(method='first').astype(int)

            punten = {}
            for i, row in enumerate(klasse_result.itertuples(), start=1):
                punten[row.bib] = i if i < 60 else 60

            for _, rijder in klasse_deelnemers.iterrows():
                bib = rijder['bib']
                punten_per_rijder.append({'bib': bib, week_col: punten.get(bib, MAX_POINTS)})
                klasse_per_rijder.append({'bib': bib, f'Klasse_{week_col}': rijder['klasse']})
                categorie_per_rijder.append({'bib': bib, f'Cat_{week_col}': rijder['categorie']})

        punten_df = pd.DataFrame(punten_per_rijder)
        klasse_df = pd.DataFrame(klasse_per_rijder)
        categorie_df = pd.DataFrame(categorie_per_rijder)

        if os.path.isfile(KLASSEMENT_FILE):
            klassement_df = pd.read_excel(KLASSEMENT_FILE, sheet_name="REGELMATIGHEIDSCRITERIUM")
            klassement_df = klassement_df.rename(columns={
                'Nr.': 'bib',
                'Naam': 'naam',
                'Klasse': 'klasse',
                'Cat.': 'categorie',
                'Totaal': 'total',
                '1e Periode': 'eerst_heft',
                '2e Periode': 'tweede_heft'
            })
        else:
            klassement_df = deelnemers[['naam', 'bib', 'klasse', 'categorie']].copy()

        klassement_df = klassement_df.merge(punten_df, on='bib', how='left')
        klassement_df = klassement_df.merge(klasse_df, on='bib', how='left')
        klassement_df = klassement_df.merge(categorie_df, on='bib', how='left')

        klassement_df[week_col] = klassement_df[week_col].fillna(MAX_POINTS)
        klassement_df[f'Klasse_{week_col}'] = klassement_df[f'Klasse_{week_col}'].fillna('Unknown')
        klassement_df[f'Cat_{week_col}'] = klassement_df[f'Cat_{week_col}'].fillna('Unknown')

        week_cols = [col for col in klassement_df.columns if col.isdigit()]
        week_cols = sorted(week_cols, key=int)

        # --- Detect klasse wissels en pas punten aan ---
        wissels = detect_klasse_wissels_met_backup()
        for bib, (oude_klasse, nieuwe_klasse) in wissels.items():
            if bib in klassement_df['bib'].values:
                idx = klassement_df.index[klassement_df['bib'] == bib][0]
                for col in week_cols:
                    if int(col) < int(week_col):
                        klassement_df.at[idx, col] = 50  # 50 punten voor oude wedstrijden

        def sum_without_worst(row, cols):
            results = row[cols].values
            if len(results) > 1:
                return results.sum() - results.max()
            else:
                return results.sum()

        klassement_df['total'] = klassement_df[week_cols].apply(lambda row: sum_without_worst(row, week_cols), axis=1)

        if week_cols:
            if IS_SECOND_PERIOD_STARTED:
                first_period_weeks = [col for col in week_cols if int(col) < int(week_cols[-1])]
                second_period_weeks = [col for col in week_cols if int(col) >= int(week_cols[-1])]
            else:
                first_period_weeks = week_cols
                second_period_weeks = []

            klassement_df['eerst_heft'] = (
                klassement_df[first_period_weeks].apply(lambda row: sum_without_worst(row, first_period_weeks), axis=1)
                if first_period_weeks else 0
            )
            klassement_df['tweede_heft'] = (
                klassement_df[second_period_weeks].apply(lambda row: sum_without_worst(row, second_period_weeks), axis=1)
                if second_period_weeks else 0
            )

        klassement_df['current_klasse'] = klassement_df[f'Klasse_{week_cols[-1]}'] if week_cols else klassement_df['klasse']
        klassement_df = klassement_df.sort_values(by=['current_klasse', 'total'])
        klassement_df['Plaats Klasse'] = (
            klassement_df.groupby('current_klasse').cumcount() + 1
        )

        for cat in ['STA', 'SEN', 'DAM']:
            klassement_df[f'Plaats {cat}'] = pd.NA
        
        klassement_df = klassement_df.reset_index(drop=True)
        
        for cat in ['STA', 'SEN', 'DAM']:
            cat_mask = klassement_df['categorie'] == cat
            if cat_mask.any():
                cat_positions = klassement_df.index[cat_mask].tolist()
                for i, pos in enumerate(cat_positions):
                    klassement_df.at[pos, f'Plaats {cat}'] = i + 1

        klassement_df = klassement_df.rename(columns={
            'bib': 'Nr.',
            'naam': 'Naam',
            'klasse': 'Klasse',
            'categorie': 'Cat.',
            'total': 'Totaal',
            'eerst_heft': '1e Periode',
            'tweede_heft': '2e Periode'
        })

        klassement_df.drop(columns=['current_klasse'], inplace=True)

        final_column_order = load_template_column_order()
        if 'Plaats Klasse' not in final_column_order:
            try:
                klasse_idx = final_column_order.index('Klasse')
                final_column_order.insert(klasse_idx + 1, 'Plaats Klasse')
            except ValueError:
                final_column_order.append('Plaats Klasse')

        for col in ['Plaats STA', 'Plaats SEN', 'Plaats DAM']:
            if col not in final_column_order:
                try:
                    cat_idx = final_column_order.index('Cat.')
                    final_column_order.insert(cat_idx + 1, col)
                except ValueError:
                    final_column_order.append(col)

        week_cols_in_output = [col for col in klassement_df.columns if col.isdigit()]
        final_cols = final_column_order + [col for col in week_cols_in_output if col not in final_column_order]
        klassement_df = klassement_df[[col for col in final_cols if col in klassement_df.columns]]

        # Save main file in output
        with pd.ExcelWriter(KLASSEMENT_FILE, engine='openpyxl', mode='w') as writer:
            klassement_df.to_excel(writer, sheet_name="REGELMATIGHEIDSCRITERIUM", index=False)

        wb = load_workbook(KLASSEMENT_FILE)
        sheet = wb['REGELMATIGHEIDSCRITERIUM']

        pink_fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        blue_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")

        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            if row[3].value == 'DAM':
                for cell in row:
                    cell.fill = pink_fill

        header = [cell.value for cell in sheet[1]]
        for idx, col_name in enumerate(header, start=1):
            if str(col_name).isdigit():
                fill = green_fill if int(col_name) <= 4 else blue_fill
                for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
                    row[idx - 1].fill = fill

        wb.save(KLASSEMENT_FILE)
        logger.info(f"✅ Week {week_num} toegevoegd aan {KLASSEMENT_FILE}")

        # --- Save backup using shared backup system ---
        backup_path = backup_file(KLASSEMENT_FILE, f"regelmatigheids_criterium_{week_num}.xlsx")
        logger.info(f"📁 Backup saved to {backup_path}")

    except Exception as e:
        logger.error(f"❌ Error in generate_regelmatigheidscriterium: {e}")
        raise

if __name__ == '__main__':
    generate_regelmatigheidscriterium()