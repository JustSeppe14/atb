import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

MAX_POINTS = 80
DEELNEMERS_FILE = "Deelnemers/deelnemerslijst 2025.xlsx"
RESULT_FILE = "Result/finish.xlsx"
KLASSEMENT_FILE = "klassement_2025.xlsx"

def load_deelnemers():
    df = pd.read_excel(DEELNEMERS_FILE, header=4)
    df.columns = df.columns.str.strip().str.lower()
    df = df.rename(columns={
        'number': 'bib',
        'name': 'naam',
        'klasse': 'klasse',
        'cat': 'categorie',
        'team': 'team'
    })
    df = df.dropna(subset=['bib', 'naam', 'klasse'])
    df['naam'] = df['naam'].str.strip().str.lower()
    df['bib'] = df['bib'].astype(int)
    return df

def load_result():
    df = pd.read_excel(RESULT_FILE)
    df.columns = df.columns.str.strip().str.lower()
    df = df.rename(columns={'pl': 'plaats', 'bib': 'bib', 'naam': 'naam'})
    df['bib'] = pd.to_numeric(df['bib'], errors='coerce').astype('Int64')
    df = df.dropna(subset=['bib', 'plaats'])
    return df

def get_current_week(overall_path):
    if not os.path.isfile(overall_path):
        return 1
    df = pd.read_excel(overall_path, sheet_name="Klassement")
    week_cols = [col for col in df.columns if str(col).isdigit()]
    return len(week_cols) + 1

def update_klassement():
    deelnemers = load_deelnemers()
    uitslag = load_result()
    week_num = get_current_week(KLASSEMENT_FILE)
    week_col = str(week_num)

    punten_per_rijder = []

    for klasse in deelnemers['klasse'].unique():
        klasse_deelnemers = deelnemers[deelnemers['klasse'] == klasse]
        klasse_bibs = set(klasse_deelnemers['bib'])

        klasse_result = uitslag[uitslag['bib'].isin(klasse_bibs)].sort_values(by='plaats')

        punten = {}
        for i, row in enumerate(klasse_result.itertuples(), start=1):
            punten[row.bib] = i if i < MAX_POINTS else MAX_POINTS

        for _, rijder in klasse_deelnemers.iterrows():
            punten_per_rijder.append({
                'bib': rijder['bib'],
                week_col: punten.get(rijder['bib'], MAX_POINTS)
            })

    punten_df = pd.DataFrame(punten_per_rijder)

    if os.path.isfile(KLASSEMENT_FILE):
        klassement_df = pd.read_excel(KLASSEMENT_FILE, sheet_name="Klassement")
    else:
        klassement_df = deelnemers[['naam', 'bib', 'klasse', 'categorie']].copy()

    klassement_df = klassement_df.merge(punten_df, on='bib', how='left')
    klassement_df[week_col] = klassement_df[week_col].fillna(MAX_POINTS)

    week_cols = [col for col in klassement_df.columns if str(col).isdigit()]
    week_cols = sorted(week_cols, key=int)

    klassement_df['total'] = klassement_df[week_cols].sum(axis=1)

    if week_cols:
        half = len(week_cols) // 2
        klassement_df['eerst_heft'] = klassement_df[week_cols[:half]].sum(axis=1)
        klassement_df['tweede_heft'] = klassement_df[week_cols[half:]].sum(axis=1)

    klassement_df = klassement_df.sort_values(by=['klasse', 'total'])

    with pd.ExcelWriter(KLASSEMENT_FILE, engine='openpyxl', mode='w') as writer:
        klassement_df.to_excel(writer, sheet_name="Klassement", index=False)

    wb = load_workbook(KLASSEMENT_FILE)
    sheet = wb['Klassement']

    pink_fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    blue_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")

    # Highlight DAM rows
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        for cell in row:
            if cell.column == 4 and cell.value == 'DAM':
                for c in row:
                    c.fill = pink_fill
                break  # No need to check further cells in this row

    # Highlight week columns
    header = [cell.value for cell in sheet[1]]
    for idx, col_name in enumerate(header, start=1):
        if str(col_name).isdigit():
            week_number = int(col_name)
            fill = green_fill if week_number <= 4 else blue_fill
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
                row[idx - 1].fill = fill

    wb.save(KLASSEMENT_FILE)

    print(f"✅ Week {week_num} toegevoegd aan {KLASSEMENT_FILE}")

if __name__ == '__main__':
    update_klassement()
