import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import logging
import shutil
from datetime import datetime

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

from utils import (
    load_deelnemers,
    load_result,
    get_current_week,
    load_template_column_order,
    detect_klasse_wissels_met_backup,
    backup_file,
    MAX_POINTS,
    DEELNEMERS_FILE,
    RESULT_FILE,
    TEMPLATE_FILE
)
OUTPUT_DIR = "output"
os.makedirs(OUTPUT_DIR, exist_ok=True)
KLASSEMENT_FILE = os.path.join(OUTPUT_DIR, "klassement_totaal_2025.xlsx")
IS_SECOND_PERIOD_STARTED = os.environ.get('IS_SECOND_PERIOD_STARTED', 'False').lower() == 'true'

def sum_best_50_percent(row, cols):
    results = row[cols].values
    if len(results) == 0:
        return 0
    
    sorted_results = sorted(results)
    
    num_to_include = max(1, len(results) // 2)
    
    return sum(sorted_results[:num_to_include])

def generate_klassement():
    try:
        deelnemers = load_deelnemers()
        uitslag = load_result()
        current_week = get_current_week(KLASSEMENT_FILE, sheet_name="KLASSEMENT")
        week_col = str(current_week)
        
        logger.info(f"Generating klassement for week {week_col}")

        deelnemers = deelnemers.rename(columns={
            'Nr.': 'bib',
            'number': 'bib',
            'Naam': 'naam',
            'Klasse': 'klasse',
            'Cat.': 'categorie'
        })

        punten_per_rijder = []
        for klasse in deelnemers['klasse'].unique():
            klasse_deelnemers = deelnemers[deelnemers['klasse'] == klasse]
            klasse_bibs = set(klasse_deelnemers['bib'])
            klasse_result = uitslag[uitslag['bib'].isin(klasse_bibs)].copy()
            klasse_result['rank_in_class'] = klasse_result['plaats'].rank(method='first').astype(int)

            punten = {}
            for row in klasse_result.itertuples():
                punten[row.bib] = row.rank_in_class if row.rank_in_class < 60 else 60

            for _, rijder in klasse_deelnemers.iterrows():
                punten_per_rijder.append({
                    'bib': rijder['bib'],
                    week_col: punten.get(rijder['bib'], MAX_POINTS)
                })

        punten_df = pd.DataFrame(punten_per_rijder)

        if os.path.isfile(KLASSEMENT_FILE):
            klassement_df = pd.read_excel(KLASSEMENT_FILE, sheet_name="KLASSEMENT")
            klassement_df = klassement_df.rename(columns={
                'Nr.': 'bib',
                'Naam': 'naam',
                'Klasse': 'klasse',
                'Cat.': 'categorie'
            })
        else:
            klassement_df = deelnemers[['naam', 'bib', 'klasse', 'categorie']].copy()

        klassement_df = klassement_df.merge(punten_df, on='bib', how='left')
        klassement_df[week_col] = klassement_df[week_col].fillna(MAX_POINTS)

        week_cols = [col for col in klassement_df.columns if str(col).isdigit()]
        week_cols = sorted(week_cols, key=int)
        klassement_df[week_cols] = klassement_df[week_cols].fillna(MAX_POINTS)

        # --- Detect klasse wissels en pas punten aan ---
        wissels = detect_klasse_wissels_met_backup()
        for bib, (oude_klasse, nieuwe_klasse) in wissels.items():
            if bib in klassement_df['bib'].values:
                idx = klassement_df.index[klassement_df['bib'] == bib][0]
                for col in week_cols:
                    if int(col) < int(week_col):
                        klassement_df.at[idx, col] = 50  # 50 punten voor oude wedstrijden

        klassement_df['Totaal'] = klassement_df[week_cols].apply(lambda row: sum_best_50_percent(row, week_cols), axis=1)

        if week_cols:
            if IS_SECOND_PERIOD_STARTED:
                second_period_start = week_cols[-1]
                first_period_weeks = [col for col in week_cols if int(col) < int(second_period_start)]
                second_period_weeks = [col for col in week_cols if int(col) >= int(second_period_start)]
            else:
                first_period_weeks = week_cols
                second_period_weeks = []

            klassement_df['1e Periode'] = (
                klassement_df[first_period_weeks].apply(lambda row: sum_best_50_percent(row, first_period_weeks), axis=1)
                if first_period_weeks else 0
            )
            klassement_df['2e Periode'] = (
                klassement_df[second_period_weeks].apply(lambda row: sum_best_50_percent(row, second_period_weeks), axis=1)
                if second_period_weeks else 0
            )

        # Calculate class rankings
        klassement_df['Plaats Klasse'] = klassement_df.groupby('klasse')['Totaal'].rank(method='min', ascending=True).astype(int)

        # Sort by class and total points first (this determines the Excel file order)
        klassement_df = klassement_df.sort_values(by=['klasse', 'Totaal'])

        # Calculate category rankings based on the order they appear in the sorted dataframe
        for cat in ['STA', 'SEN', 'DAM']:
            klassement_df[f'Plaats {cat}'] = pd.NA
        
        klassement_df = klassement_df.reset_index(drop=True)
        
        for cat in ['STA', 'SEN', 'DAM']:
            cat_mask = klassement_df['categorie'] == cat
            if cat_mask.any():
                cat_positions = klassement_df.index[cat_mask].tolist()
                for i, pos in enumerate(cat_positions):
                    klassement_df.at[pos, f'Plaats {cat}'] = i + 1

        klassement_df = klassement_df.sort_values(by=['klasse', 'Totaal'])

        klassement_df = klassement_df.rename(columns={
            'bib': 'Nr.',
            'naam': 'Naam',
            'klasse': 'Klasse',
            'categorie': 'Cat.'
        })
        
        if 'bib' in klassement_df.columns:
            klassement_df = klassement_df.drop('bib', axis=1)

        final_column_order = load_template_column_order()
        if 'Plaats Klasse' not in final_column_order:
            try:
                klasse_idx = final_column_order.index('Klasse')
                final_column_order.insert(klasse_idx + 1, 'Plaats Klasse')
            except ValueError:
                final_column_order.append('Plaats Klasse')
                
        for col in ['Plaats STA', 'Plaats SEN', 'Plaats DAM']:
            if col not in final_column_order:
                try:
                    cat_idx = final_column_order.index('Cat.')
                    final_column_order.insert(cat_idx + 1, col)
                except ValueError:
                    final_column_order.append(col)

        week_cols_in_output = [col for col in klassement_df.columns if col.isdigit()]
        final_cols = final_column_order + [col for col in week_cols_in_output if col not in final_column_order]
        klassement_df = klassement_df[[col for col in final_cols if col in klassement_df.columns]]

        with pd.ExcelWriter(KLASSEMENT_FILE, engine='openpyxl', mode='w') as writer:
            klassement_df.to_excel(writer, sheet_name="KLASSEMENT", index=False)

        wb = load_workbook(KLASSEMENT_FILE)
        sheet = wb['KLASSEMENT']

        pink_fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        blue_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")

        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            if row[3].value == 'DAM':
                for cell in row:
                    cell.fill = pink_fill

        header = [cell.value for cell in sheet[1]]
        for idx, col_name in enumerate(header, start=1):
            if str(col_name).isdigit():
                fill = green_fill if int(col_name) <= 4 else blue_fill
                for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
                    row[idx - 1].fill = fill

        wb.save(KLASSEMENT_FILE)
        logger.info(f"✅ Klassement updated with week {current_week} in {KLASSEMENT_FILE}")

        # --- Save backup using shared backup system ---
        backup_path = backup_file(KLASSEMENT_FILE, f"klassement_totaal_2025_week_{current_week}.xlsx")
        logger.info(f"📁 Backup saved to {backup_path}")

    except Exception as e:
        logger.error(f"❌ Error in generate_klassement: {e}")
        raise

if __name__ == '__main__':
    generate_klassement()